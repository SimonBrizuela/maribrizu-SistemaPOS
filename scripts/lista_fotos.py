"""
Arma la lista de trabajo para conseguir las fotos de la tienda.

Por que no se bajan solas: de los productos con stock, el 39% no tiene marca
cargada, y los que si la tienen se reparten entre mas de 300 proveedores, en su
mayoria mayoristas chicos de merceria que no publican catalogo de imagenes.
Buscarlas es trabajo de compras, no un script.

Lo que si se puede automatizar es decidir POR CUALES EMPEZAR. Fotografiar 2.315
productos en orden alfabetico es un mes de trabajo; fotografiar los 100 que
mueven la mitad de las ventas es una tarde, y el cliente ve fotos justo donde
mira.

Genera un Excel con dos hojas:
  · "Sacar foto"     productos sin marca, ordenados por lo que se vendieron.
  · "Pedir al prov"  productos con marca, agrupados para pedirselas al proveedor.
Mas una hoja de proveedores con el volumen de cada uno, para saber a quien
llamar primero.

    python scripts/lista_fotos.py
    python scripts/lista_fotos.py --meses 6
"""
import argparse
import os
import sys
import unicodedata
from collections import defaultdict
from datetime import datetime, timedelta, timezone

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import firebase_admin
from firebase_admin import credentials, firestore
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

RAIZ = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))


def conectar():
    try:
        firebase_admin.get_app()
    except ValueError:
        firebase_admin.initialize_app(
            credentials.Certificate(os.path.join(RAIZ, 'firebase_key.json')))
    return firestore.client()


def normalizar(texto):
    s = unicodedata.normalize('NFD', str(texto or '').lower())
    return ' '.join(''.join(c for c in s if unicodedata.category(c) != 'Mn').split())


def encabezar(hoja, columnas):
    relleno = PatternFill('solid', fgColor='0E0D10')
    for i, (titulo, ancho) in enumerate(columnas, start=1):
        celda = hoja.cell(row=1, column=i, value=titulo)
        celda.font = Font(bold=True, color='FFFFFF')
        celda.fill = relleno
        celda.alignment = Alignment(vertical='center')
        hoja.column_dimensions[get_column_letter(i)].width = ancho
    hoja.freeze_panes = 'A2'


def main():
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument('--meses', type=int, default=4,
                    help='cuantos meses de ventas mirar (por defecto 4)')
    ap.add_argument('--salida', default='FOTOS_PENDIENTES.xlsx')
    args = ap.parse_args()

    db = conectar()

    # ── Que se vendio ─────────────────────────────────────────────────────
    desde = datetime.now(timezone.utc) - timedelta(days=args.meses * 30)
    print(f'Leyendo ventas desde {desde:%d/%m/%Y}...')

    unidades = defaultdict(float)
    importe = defaultdict(float)

    consulta = db.collection('ventas_por_dia').where('fecha_dt', '>=', desde)
    leidos = 0
    for d in consulta.stream():
        x = d.to_dict() or {}
        leidos += 1
        if x.get('deleted') is True:
            continue
        nombre = normalizar(x.get('producto') or x.get('product_name'))
        if not nombre:
            continue
        unidades[nombre] += float(x.get('cantidad') or x.get('quantity') or 0)
        importe[nombre] += float(x.get('subtotal') or 0)

    print(f'  {leidos} items de venta, {len(unidades)} productos distintos\n')

    # ── Que hay publicado y sin foto ──────────────────────────────────────
    print('Leyendo el catalogo publicado...')
    # El nombre en las ventas es el del catalogo (en mayusculas), no el que se
    # muestra en la tienda, asi que el cruce se hace contra `catalogo`.
    nombre_original = {}
    for d in db.collection('catalogo').select(['nombre']).stream():
        nombre_original[d.id] = (d.to_dict() or {}).get('nombre') or ''

    filas = []
    for d in db.collection('tienda_productos').stream():
        x = d.to_dict() or {}
        if x.get('imagenes'):
            continue  # ya tiene foto
        clave = normalizar(nombre_original.get(d.id, x.get('nombre')))
        filas.append({
            'codigo': x.get('codigo') or d.id,
            'nombre': x.get('nombre') or '',
            'marca': (x.get('marca') or '').strip(),
            'rubro': x.get('rubro') or '',
            'sub_rubro': x.get('sub_rubro') or '',
            'stock': int(x.get('stock') or 0),
            'precio': int(x.get('precio') or 0),
            'unidades': round(unidades.get(clave, 0)),
            'importe': round(importe.get(clave, 0)),
        })

    # Se ordena por plata vendida, no por unidades: cien gomas de borrar mueven
    # menos que diez mochilas, y en la tienda pesa mas la mochila.
    filas.sort(key=lambda f: (-f['importe'], -f['unidades'], f['nombre']))

    sin_marca = [f for f in filas if not f['marca']]
    con_marca = [f for f in filas if f['marca']]

    print(f'  {len(filas)} productos publicados sin foto')
    print(f'    {len(sin_marca)} sin marca  -> sacarles foto')
    print(f'    {len(con_marca)} con marca  -> pedirselas al proveedor\n')

    # ── Excel ─────────────────────────────────────────────────────────────
    libro = Workbook()

    COLUMNAS = [('#', 6), ('Codigo', 14), ('Producto', 52), ('Rubro', 14),
                ('Sub rubro', 22), ('Stock', 8), ('Precio', 10),
                ('Vendidos', 10), ('Facturado', 13), ('Foto lista', 11)]

    def volcar(hoja, datos, con_columna_marca=False):
        cols = COLUMNAS[:]
        if con_columna_marca:
            cols = COLUMNAS[:3] + [('Marca', 20)] + COLUMNAS[3:]
        encabezar(hoja, cols)

        for i, f in enumerate(datos, start=1):
            fila = [i, f['codigo'], f['nombre']]
            if con_columna_marca:
                fila.append(f['marca'])
            fila += [f['rubro'].title(), f['sub_rubro'], f['stock'], f['precio'],
                     f['unidades'], f['importe'], '']
            hoja.append(fila)

            # Los que ya se vendieron van resaltados: son los que el cliente va
            # a mirar primero en la tienda.
            if f['importe'] > 0:
                for c in range(1, len(cols) + 1):
                    hoja.cell(row=i + 1, column=c).fill = PatternFill(
                        'solid', fgColor='FFF6E0')

    hoja = libro.active
    hoja.title = 'Sacar foto'
    volcar(hoja, sin_marca)

    volcar(libro.create_sheet('Pedir al proveedor'), con_marca, con_columna_marca=True)

    # ── Proveedores ───────────────────────────────────────────────────────
    por_marca = defaultdict(lambda: {'productos': 0, 'importe': 0, 'unidades': 0})
    for f in con_marca:
        m = por_marca[f['marca']]
        m['productos'] += 1
        m['importe'] += f['importe']
        m['unidades'] += f['unidades']

    hoja = libro.create_sheet('Proveedores')
    encabezar(hoja, [('Marca', 30), ('Productos sin foto', 18),
                     ('Facturado', 14), ('Vendidos', 12)])
    for marca, m in sorted(por_marca.items(), key=lambda x: -x[1]['importe']):
        hoja.append([marca, m['productos'], m['importe'], m['unidades']])

    salida = os.path.join(RAIZ, args.salida)
    libro.save(salida)

    print(f'Listo: {salida}\n')
    print('Los 15 primeros para sacarles foto:')
    for f in sin_marca[:15]:
        print(f'  ${f["importe"]:>9,d}  {f["unidades"]:>4d}u  {f["nombre"][:56]}')

    print('\nLos 10 proveedores a los que conviene pedirles primero:')
    for marca, m in sorted(por_marca.items(), key=lambda x: -x[1]['importe'])[:10]:
        print(f'  ${m["importe"]:>9,d}  {m["productos"]:>4d} productos  {marca}')


if __name__ == '__main__':
    main()
